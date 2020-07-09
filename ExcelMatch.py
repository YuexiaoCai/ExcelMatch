from openpyxl import *
import re

# 打开excel文件,获取工作簿对象
wb_input = load_workbook(filename='南航-计算机科学与技术学院.xlsx')
wb_output = load_workbook(filename='StudentList.xlsx')

#获取当前表单
ws_input = wb_input.active
ws_output = wb_output.active

#初始化参与学习的同学列表，规定每项格式[学号,姓名]
student = []
temp = []

#正则表达式
str_id = re.compile(r'\d{9}')
str_name = re.compile(r'[\u4e00-\u9fa5]{2,9}')

#读取参与的同学信息
for index in range(2, ws_input.max_row + 1):
    #获取正则匹配信息
    id_match = str_id.search(str(ws_input['A'+str(index)].value))
    name_match = str_name.search(str(ws_input['A'+str(index)].value))
    #如果匹配失败则添加空信息
    if id_match == None:
        temp.append('')
    else:
        temp.append(id_match.group())
    if name_match == None:
        temp.append('')
    else:
        temp.append(name_match.group())
    #将同学信息添加至列表，重置临时列表
    student.append(temp)
    temp=[]

#根据学号或姓名判断是否学习
ws_output['E2'] = '是否学习'
for index in range(3, ws_output.max_row + 1):
    id = str(ws_output['B'+str(index)].value)
    flag = 0
    for i in student:
        if id == i[0]:
            flag = 1
            break
    if flag:
        ws_output['E'+str(index)] = 1
    else:
        ws_output['E'+str(index)] = 0

for index in range(3, ws_output.max_row + 1):
    id = str(ws_output['C'+str(index)].value)
    flag = 0
    for i in student:
        if id == i[1]:
            flag = 1
            break
    if flag:
        ws_output['E'+str(index)] = 1
    

#根据学号姓名统计班级学习人数
#新建工作簿，添加表头信息
wb_output_class = Workbook()
ws_output_class = wb_output_class.active
ws_output_class['A1'] = '班级'
ws_output_class['B1'] = '学习人数'
#初始化班级人数变量
class_name = ws_output['A3'].value
count_id = 0
count_name = 0
count_all = 0
count_or = 0 
count = 2
for index in range(3, ws_output.max_row + 2):
    #获取新班级
    if ws_output['A'+str(index)].value != class_name :
        ws_output_class['A' + str(count)] = class_name
        ws_output_class['B' + str(count)] = count_or
        count += 1
        class_name = ws_output['A'+str(index)].value
        count_id = 0
        count_name = 0
        count_all = 0
        count_or = 0 
    #增加班级学习人数
    if ws_output['E'+str(index)].value == 1:
        count_or += 1

#另存为新文件
wb_output.save('学习名单.xlsx')
wb_output_class.save('班级情况.xlsx')
